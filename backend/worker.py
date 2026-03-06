from __future__ import annotations

import csv
import io
import json
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import redis
from celery import Task
from openpyxl import Workbook, load_workbook
from pypdf import PdfReader, PdfWriter

from celery_app import celery_app
from config import settings
from database import SessionLocal, init_db
from models import DataRecord, DesensitizeTask, TaskStatus

# Initialize Redis client for Pub/Sub
redis_client = redis.from_url(settings.celery_broker_url)

EMAIL_KEYWORDS = {"email", "e-mail", "mail", "邮箱"}
PHONE_KEYWORDS = {"phone", "mobile", "tel", "telephone", "手机号", "电话"}
ID_KEYWORDS = {"id_card", "idcard", "identity", "ssn", "passport", "身份证", "证件"}
NAME_KEYWORDS = {"name", "full_name", "first_name", "last_name", "姓名"}
ADDRESS_KEYWORDS = {"address", "addr", "地址"}
DEFAULT_SPLIT_CHUNK_BYTES = 140 * 1024 * 1024


def _iter_csv_rows(path: Path):
    with path.open("r", newline="") as handle:
        reader = csv.reader(handle)
        yield from reader


def _iter_xlsx_rows(path: Path):
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            yield row
    finally:
        workbook.close()


def _iter_json_rows(path: Path):
    with path.open("r", encoding="utf-8") as handle:
        data = json.load(handle)
    if isinstance(data, list):
        for item in data:
            yield list(item.values()) if isinstance(item, dict) else [item]
    elif isinstance(data, dict):
        yield list(data.keys())
        yield list(data.values())
    else:
        yield [data]


def _iter_jsonl_rows(path: Path):
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if line:
                item = json.loads(line)
                yield list(item.values()) if isinstance(item, dict) else [item]


def _count_rows(path: Path, file_type: str) -> int:
    if file_type == "csv":
        return sum(1 for _ in _iter_csv_rows(path))
    if file_type == "xlsx":
        return sum(1 for _ in _iter_xlsx_rows(path))
    if file_type == "json":
        return sum(1 for _ in _iter_json_rows(path))
    if file_type == "jsonl":
        return sum(1 for _ in _iter_jsonl_rows(path))
    raise ValueError(f"unsupported file type: {file_type}")


def _iter_rows(path: Path, file_type: str):
    if file_type == "csv":
        return _iter_csv_rows(path)
    if file_type == "xlsx":
        return _iter_xlsx_rows(path)
    if file_type == "json":
        return _iter_json_rows(path)
    if file_type == "jsonl":
        return _iter_jsonl_rows(path)
    raise ValueError(f"unsupported file type: {file_type}")


def _row_payload(row) -> str:
    return ",".join("" if cell is None else str(cell) for cell in row)


def _select_masker(header: str):
    normalized = header.strip().lower()
    if any(key in normalized for key in EMAIL_KEYWORDS):
        return _mask_email
    if any(key in normalized for key in PHONE_KEYWORDS):
        return _mask_phone
    if any(key in normalized for key in ID_KEYWORDS):
        return _mask_id
    if any(key in normalized for key in NAME_KEYWORDS):
        return _mask_name
    if any(key in normalized for key in ADDRESS_KEYWORDS):
        return _mask_address
    return None


def _mask_email(value: str) -> str:
    if "@" not in value:
        return _mask_generic(value)
    local, domain = value.split("@", 1)
    if not local:
        return f"***@{domain}"
    return f"{local[0]}***@{domain}"


def _mask_phone(value: str) -> str:
    digits = [char for char in value if char.isdigit()]
    if not digits:
        return _mask_generic(value)
    keep = 4 if len(digits) > 4 else 0
    masked = []
    seen = 0
    for char in value:
        if char.isdigit():
            seen += 1
            if len(digits) - seen < keep:
                masked.append(char)
            else:
                masked.append("*")
        else:
            masked.append(char)
    return "".join(masked)


def _mask_id(value: str) -> str:
    if len(value) <= 4:
        return "*" * len(value)
    return f"{value[:2]}{'*' * (len(value) - 4)}{value[-2:]}"


def _mask_name(value: str) -> str:
    if len(value) <= 1:
        return "*" * len(value)
    return f"{value[0]}{'*' * (len(value) - 1)}"


def _mask_address(value: str) -> str:
    if len(value) <= 6:
        return "*" * len(value)
    return f"{value[:6]}{'*' * (len(value) - 6)}"


def _mask_generic(value: str) -> str:
    if len(value) <= 2:
        return "*" * len(value)
    return f"{value[0]}{'*' * (len(value) - 2)}{value[-1]}"


def _apply_mask(value, masker):
    if value is None:
        return ""
    text = str(value)
    if not text:
        return text
    return masker(text) if masker else text


def _part_path(output_dir: Path, task_id: str, part_index: int, suffix: str) -> Path:
    return output_dir / f"{task_id}_part_{part_index:03d}{suffix}"


def _write_pdf_pages(pages, destination: Path) -> int:
    writer = PdfWriter()
    for page in pages:
        writer.add_page(page)
    with destination.open("wb") as handle:
        writer.write(handle)
    return destination.stat().st_size


def _pdf_pages_size(pages) -> int:
    writer = PdfWriter()
    for page in pages:
        writer.add_page(page)
    buffer = io.BytesIO()
    writer.write(buffer)
    return buffer.tell()


def _split_txt_parts(
    source_path: Path,
    output_dir: Path,
    task_id: str,
    chunk_size_bytes: int,
) -> list[Path]:
    if chunk_size_bytes <= 0:
        raise ValueError("chunk_size_bytes must be greater than 0")

    part_paths: list[Path] = []
    part_index = 1
    current_path = _part_path(output_dir, task_id, part_index, ".txt")
    current_file = current_path.open("wb")
    current_size = 0
    part_paths.append(current_path)

    try:
        with source_path.open("rb") as source:
            while True:
                chunk = source.read(1024 * 1024)
                if not chunk:
                    break

                offset = 0
                chunk_length = len(chunk)
                while offset < chunk_length:
                    available = chunk_size_bytes - current_size
                    if available == 0:
                        current_file.close()
                        part_index += 1
                        current_path = _part_path(
                            output_dir, task_id, part_index, ".txt"
                        )
                        part_paths.append(current_path)
                        current_file = current_path.open("wb")
                        current_size = 0
                        available = chunk_size_bytes

                    take = min(available, chunk_length - offset)
                    current_file.write(chunk[offset : offset + take])
                    offset += take
                    current_size += take
    finally:
        current_file.close()

    return part_paths


def _split_pdf_parts(
    source_path: Path,
    output_dir: Path,
    task_id: str,
    chunk_size_bytes: int,
) -> list[Path]:
    if chunk_size_bytes <= 0:
        raise ValueError("chunk_size_bytes must be greater than 0")

    reader = PdfReader(str(source_path))
    pages = list(reader.pages)
    if not pages:
        output_path = _part_path(output_dir, task_id, 1, ".pdf")
        _write_pdf_pages([], output_path)
        return [output_path]

    part_paths: list[Path] = []
    current_pages = []
    part_index = 1

    for page in pages:
        current_pages.append(page)
        if _pdf_pages_size(current_pages) <= chunk_size_bytes:
            continue

        if len(current_pages) == 1:
            output_path = _part_path(output_dir, task_id, part_index, ".pdf")
            _write_pdf_pages(current_pages, output_path)
            part_paths.append(output_path)
            part_index += 1
            current_pages = []
            continue

        overflow_page = current_pages.pop()
        output_path = _part_path(output_dir, task_id, part_index, ".pdf")
        _write_pdf_pages(current_pages, output_path)
        part_paths.append(output_path)
        part_index += 1

        current_pages = [overflow_page]
        if _pdf_pages_size(current_pages) > chunk_size_bytes:
            output_path = _part_path(output_dir, task_id, part_index, ".pdf")
            _write_pdf_pages(current_pages, output_path)
            part_paths.append(output_path)
            part_index += 1
            current_pages = []

    if current_pages:
        output_path = _part_path(output_dir, task_id, part_index, ".pdf")
        _write_pdf_pages(current_pages, output_path)
        part_paths.append(output_path)

    return part_paths


def split_file_and_build_zip(
    source_path: Path,
    output_dir: Path,
    task_id: str,
    chunk_size_bytes: int = DEFAULT_SPLIT_CHUNK_BYTES,
) -> tuple[Path, list[Path]]:
    output_dir.mkdir(parents=True, exist_ok=True)

    suffix = source_path.suffix.lower()
    if suffix == ".txt":
        part_paths = _split_txt_parts(
            source_path, output_dir, task_id, chunk_size_bytes
        )
    elif suffix == ".pdf":
        part_paths = _split_pdf_parts(
            source_path, output_dir, task_id, chunk_size_bytes
        )
    else:
        raise ValueError(f"unsupported file type: {suffix}")

    zip_path = output_dir / f"{task_id}_split.zip"
    with ZipFile(zip_path, "w", compression=ZIP_DEFLATED) as archive:
        for part_path in part_paths:
            archive.write(part_path, arcname=part_path.name)

    return zip_path, part_paths


@celery_app.task(bind=True)
def process_csv(self: Task, file_path: str) -> dict:
    init_db()
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"file not found: {file_path}")

    suffix = path.suffix.lower()
    if suffix not in {".csv", ".xlsx"}:
        raise ValueError(f"unsupported file type: {suffix}")
    file_type = "xlsx" if suffix == ".xlsx" else "csv"

    total = _count_rows(path, file_type)

    if total == 0:
        return {"current": 0, "total": 0, "message": "no rows"}

    session = SessionLocal()
    try:
        for index, row in enumerate(_iter_rows(path, file_type), start=1):
            record = DataRecord(
                task_id=self.request.id,
                row_number=index,
                payload=_row_payload(row),
            )
            session.add(record)

            if index % 100 == 0:
                session.commit()

            if index == total or index % 10 == 0:
                self.update_state(
                    state="PROGRESS",
                    meta={
                        "current": index,
                        "total": total,
                        "message": f"Processing row {index}/{total}",
                    },
                )
                # Publish progress to Redis
                redis_client.publish(
                    f"task_progress:{self.request.id}",
                    json.dumps(
                        {
                            "current": index,
                            "total": total,
                            "message": f"Processing row {index}/{total}",
                        }
                    ),
                )

        session.commit()
    finally:
        session.close()

    # Publish completion message
    redis_client.publish(
        f"task_progress:{self.request.id}",
        json.dumps({"current": total, "total": total, "message": "completed"}),
    )

    return {"current": total, "total": total, "message": "completed"}


@celery_app.task(bind=True)
def process_desensitize(self: Task, file_path: str) -> dict:
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"file not found: {file_path}")

    suffix = path.suffix.lower()
    if suffix not in {".csv", ".xlsx"}:
        raise ValueError(f"unsupported file type: {suffix}")
    file_type = "xlsx" if suffix == ".xlsx" else "csv"

    output_dir = Path(settings.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{self.request.id}_desensitized{suffix}"

    total_rows = _count_rows(path, file_type)
    rows = _iter_rows(path, file_type)
    try:
        header = next(rows)
    except StopIteration:
        return {"current": 0, "total": 0, "message": "no rows"}

    header_cells = ["" if cell is None else str(cell) for cell in header]
    maskers = [_select_masker(cell) for cell in header_cells]

    data_total = max(total_rows - 1, 0)

    if file_type == "csv":
        with output_path.open("w", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(header_cells)
            for index, row in enumerate(rows, start=1):
                masked_row = []
                for i, cell in enumerate(row):
                    masker = maskers[i] if i < len(maskers) else None
                    masked_row.append(_apply_mask(cell, masker))
                writer.writerow(masked_row)

                if data_total and (index == data_total or index % 10 == 0):
                    message = f"Desensitizing row {index}/{data_total}"
                    self.update_state(
                        state="PROGRESS",
                        meta={
                            "current": index,
                            "total": data_total,
                            "message": message,
                        },
                    )
                    redis_client.publish(
                        f"task_progress:{self.request.id}",
                        json.dumps(
                            {"current": index, "total": data_total, "message": message}
                        ),
                    )
    elif file_type == "xlsx":
        workbook = Workbook(write_only=True)
        sheet = workbook.create_sheet()
        sheet.append(header_cells)
        for index, row in enumerate(rows, start=1):
            masked_row = []
            for i, cell in enumerate(row):
                masker = maskers[i] if i < len(maskers) else None
                masked_row.append(_apply_mask(cell, masker))
            sheet.append(masked_row)

            if data_total and (index == data_total or index % 10 == 0):
                message = f"Desensitizing row {index}/{data_total}"
                self.update_state(
                    state="PROGRESS",
                    meta={"current": index, "total": data_total, "message": message},
                )
                redis_client.publish(
                    f"task_progress:{self.request.id}",
                    json.dumps(
                        {"current": index, "total": data_total, "message": message}
                    ),
                )
        workbook.save(output_path)
        workbook.close()
    elif file_type in ("json", "jsonl"):
        rows_list = []
        rows_iter = _iter_rows(path, file_type)
        try:
            header = next(rows_iter)
        except StopIteration:
            header = []
        header_cells = ["" if cell is None else str(cell) for cell in header]

        for index, row in enumerate(rows_iter, start=1):
            masked_row = []
            for i, cell in enumerate(row):
                masker = maskers[i] if i < len(maskers) else None
                masked_row.append(_apply_mask(cell, masker))
            row_dict = dict(zip(header_cells, masked_row))
            rows_list.append(row_dict)

            if data_total and (index == data_total or index % 10 == 0):
                message = f"Desensitizing row {index}/{data_total}"
                self.update_state(
                    state="PROGRESS",
                    meta={"current": index, "total": data_total, "message": message},
                )
                redis_client.publish(
                    f"task_progress:{self.request.id}",
                    json.dumps(
                        {"current": index, "total": data_total, "message": message}
                    ),
                )

        with output_path.open("w", encoding="utf-8") as handle:
            if file_type == "json":
                json.dump(rows_list, handle, ensure_ascii=False, indent=2)
            else:
                for row in rows_list:
                    handle.write(json.dumps(row, ensure_ascii=False) + "\n")

    if data_total == 0:
        redis_client.publish(
            f"task_progress:{self.request.id}",
            json.dumps({"current": 0, "total": 0, "message": "completed"}),
        )
        return {
            "current": 0,
            "total": 0,
            "message": "completed",
            "output_file": output_path.name,
        }

    redis_client.publish(
        f"task_progress:{self.request.id}",
        json.dumps(
            {"current": data_total, "total": data_total, "message": "completed"}
        ),
    )

    return {
        "current": data_total,
        "total": data_total,
        "message": "completed",
        "output_file": output_path.name,
    }


def _update_task_record(
    task_db_id: str,
    *,
    status: TaskStatus | None = None,
    progress: float | None = None,
    message: str | None = None,
    error_detail: str | None = None,
    celery_task_id: str | None = None,
    input_rows: int | None = None,
    output_rows: int | None = None,
    masked_fields: int | None = None,
    started_at=None,
    completed_at=None,
) -> None:
    """Update a DesensitizeTask record in the database."""
    session = SessionLocal()
    try:
        task_record = (
            session.query(DesensitizeTask)
            .filter(DesensitizeTask.id == task_db_id)
            .first()
        )
        if not task_record:
            return
        if status is not None:
            task_record.status = status
        if progress is not None:
            task_record.progress = progress
        if message is not None:
            task_record.message = message
        if error_detail is not None:
            task_record.error_detail = error_detail
        if celery_task_id is not None:
            task_record.celery_task_id = celery_task_id
        if input_rows is not None:
            task_record.input_rows = input_rows
        if output_rows is not None:
            task_record.output_rows = output_rows
        if masked_fields is not None:
            task_record.masked_fields = masked_fields
        if started_at is not None:
            task_record.started_at = started_at
        if completed_at is not None:
            task_record.completed_at = completed_at
        session.commit()
    finally:
        session.close()


@celery_app.task(bind=True)
def process_split_archive(self: Task, file_path: str, chunk_size_mb: int = 140) -> dict:
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"file not found: {file_path}")

    suffix = path.suffix.lower()
    if suffix not in {".pdf", ".txt"}:
        raise ValueError(f"unsupported file type: {suffix}")

    output_dir = Path(settings.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    chunk_size_bytes = chunk_size_mb * 1024 * 1024
    self.update_state(
        state="PROGRESS",
        meta={"current": 0, "total": 1, "message": "splitting file"},
    )
    redis_client.publish(
        f"task_progress:{self.request.id}",
        json.dumps({"current": 0, "total": 1, "message": "splitting file"}),
    )

    zip_path, part_paths = split_file_and_build_zip(
        path,
        output_dir,
        task_id=self.request.id,
        chunk_size_bytes=chunk_size_bytes,
    )

    part_count = len(part_paths)
    message = f"completed ({part_count} parts)"
    redis_client.publish(
        f"task_progress:{self.request.id}",
        json.dumps({"current": part_count, "total": part_count, "message": message}),
    )

    return {
        "current": part_count,
        "total": part_count,
        "message": message,
        "output_file": zip_path.name,
    }


@celery_app.task(bind=True)
def process_db_desensitize(self: Task, task_db_id: str) -> dict:
    """
    Database-to-CSV desensitization task.

    Reads data from a source database table, applies masking rules,
    and writes the masked output to a CSV file.
    """
    from datetime import datetime

    from connectors.factory import create_connector
    from discovery.scanner import scan_table_schema
    from masking.engine import MaskingEngine

    init_db()

    # Load task record from DB
    session = SessionLocal()
    try:
        task_record = (
            session.query(DesensitizeTask)
            .filter(DesensitizeTask.id == task_db_id)
            .first()
        )
        if not task_record:
            raise ValueError(f"DesensitizeTask not found: {task_db_id}")

        source_config = task_record.source_config or {}
        target_config = task_record.target_config or {}
        rules_config = task_record.rules or {}
    finally:
        session.close()

    # Mark task as running
    _update_task_record(
        task_db_id,
        status=TaskStatus.RUNNING,
        celery_task_id=self.request.id,
        started_at=datetime.utcnow(),
    )

    try:
        # Determine connector type from source_config
        connector_type = source_config.get("connector_type", "")
        if not connector_type:
            raise ValueError("source_config must include 'connector_type'")

        table_name = source_config.get("table")
        if not table_name:
            raise ValueError("source_config must include 'table'")

        connection_config = source_config.get("connection", {})

        # Connect to source database
        connector = create_connector(connector_type, connection_config)
        connector.connect()

        try:
            # Build masking engine
            rules_list = rules_config.get("rules", [])
            if rules_list:
                # Use explicitly provided rules
                engine = MaskingEngine.from_rules_config(rules_list)
            else:
                # Auto-discover sensitive columns using scanner
                columns = connector.get_columns(table_name)
                samples = connector.get_sample_data(table_name, limit=5)
                sample_values: dict[str, str | None] = {}
                if samples:
                    first_row = samples[0]
                    for col_name, val in first_row.items():
                        sample_values[col_name] = str(val) if val is not None else None

                scan_result = scan_table_schema(table_name, columns, sample_values)
                # Convert discovered fields into masking rules
                auto_rules = []
                for field in scan_result.fields:
                    auto_rules.append(
                        {
                            "column_name": field.column_name,
                            "masking_type": field.data_type.value,
                        }
                    )
                engine = MaskingEngine.from_rules_config(auto_rules)

            masked_field_count = len(engine.rules)

            # Query all rows from source table
            query = source_config.get("query")
            if query:
                rows = connector.execute_query(query)
            else:
                rows = connector.execute_query(f"SELECT * FROM {table_name}")

            total_rows = len(rows)

            # Prepare output
            output_dir = Path(target_config.get("output_dir", settings.output_dir))
            output_dir.mkdir(parents=True, exist_ok=True)
            output_path = output_dir / f"{task_db_id}_desensitized.csv"

            # Publish initial progress
            self.update_state(
                state="PROGRESS",
                meta={
                    "current": 0,
                    "total": total_rows,
                    "message": "Starting desensitization",
                },
            )
            redis_client.publish(
                f"task_progress:{self.request.id}",
                json.dumps(
                    {
                        "current": 0,
                        "total": total_rows,
                        "message": "Starting desensitization",
                    }
                ),
            )

            if total_rows == 0:
                # Write empty CSV with headers if available
                with output_path.open("w", newline="") as handle:
                    writer = csv.writer(handle)
                    # Try to get column names from schema
                    columns = connector.get_columns(table_name)
                    if columns:
                        writer.writerow([col["name"] for col in columns])

                _update_task_record(
                    task_db_id,
                    status=TaskStatus.COMPLETED,
                    progress=1.0,
                    message="completed (no rows)",
                    input_rows=0,
                    output_rows=0,
                    masked_fields=masked_field_count,
                    completed_at=datetime.utcnow(),
                )

                redis_client.publish(
                    f"task_progress:{self.request.id}",
                    json.dumps({"current": 0, "total": 0, "message": "completed"}),
                )

                return {
                    "current": 0,
                    "total": 0,
                    "message": "completed",
                    "output_file": output_path.name,
                    "input_rows": 0,
                    "output_rows": 0,
                    "masked_fields": masked_field_count,
                }

            # Write masked rows to CSV
            header = list(rows[0].keys())
            with output_path.open("w", newline="") as handle:
                writer = csv.writer(handle)
                writer.writerow(header)

                for index, row in enumerate(rows, start=1):
                    masked_row = engine.apply_to_row(row)
                    writer.writerow([masked_row.get(col, "") for col in header])

                    if index == total_rows or index % 50 == 0:
                        progress = index / total_rows
                        message = f"Desensitizing row {index}/{total_rows}"
                        self.update_state(
                            state="PROGRESS",
                            meta={
                                "current": index,
                                "total": total_rows,
                                "message": message,
                            },
                        )
                        redis_client.publish(
                            f"task_progress:{self.request.id}",
                            json.dumps(
                                {
                                    "current": index,
                                    "total": total_rows,
                                    "message": message,
                                }
                            ),
                        )
                        _update_task_record(
                            task_db_id,
                            progress=progress,
                            message=message,
                        )

        finally:
            connector.disconnect()

        # Mark completed
        _update_task_record(
            task_db_id,
            status=TaskStatus.COMPLETED,
            progress=1.0,
            message="completed",
            input_rows=total_rows,
            output_rows=total_rows,
            masked_fields=masked_field_count,
            completed_at=datetime.utcnow(),
        )

        redis_client.publish(
            f"task_progress:{self.request.id}",
            json.dumps(
                {"current": total_rows, "total": total_rows, "message": "completed"}
            ),
        )

        return {
            "current": total_rows,
            "total": total_rows,
            "message": "completed",
            "output_file": output_path.name,
            "input_rows": total_rows,
            "output_rows": total_rows,
            "masked_fields": masked_field_count,
        }

    except Exception as exc:
        _update_task_record(
            task_db_id,
            status=TaskStatus.FAILED,
            error_detail=str(exc),
        )
        redis_client.publish(
            f"task_progress:{self.request.id}",
            json.dumps({"current": 0, "total": 0, "message": f"failed: {exc}"}),
        )
        raise


@celery_app.task(bind=True)
def process_task_desensitize(self: Task, task_db_id: str) -> dict:
    """
    File-based desensitization task dispatched through the task API.

    Reads the file path from DesensitizeTask.source_config, applies
    desensitization using the same logic as process_desensitize, and
    updates the DesensitizeTask DB record with progress/completion/error.
    """
    from datetime import datetime

    init_db()

    # Load task record from DB
    session = SessionLocal()
    try:
        task_record = (
            session.query(DesensitizeTask)
            .filter(DesensitizeTask.id == task_db_id)
            .first()
        )
        if not task_record:
            raise ValueError(f"DesensitizeTask not found: {task_db_id}")

        source_config = task_record.source_config or {}
        target_config = task_record.target_config or {}
        rules_config = task_record.rules or {}
    finally:
        session.close()

    # Mark task as running
    _update_task_record(
        task_db_id,
        status=TaskStatus.RUNNING,
        celery_task_id=self.request.id,
        started_at=datetime.utcnow(),
    )

    try:
        file_path = source_config.get("file_path", "")
        if not file_path:
            raise ValueError("source_config must include 'file_path'")

        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"file not found: {file_path}")

        suffix = path.suffix.lower()
        if suffix not in {".csv", ".xlsx", ".json", ".jsonl"}:
            raise ValueError(f"unsupported file type: {suffix}")
        file_type = (
            "xlsx"
            if suffix == ".xlsx"
            else "json"
            if suffix == ".json"
            else "jsonl"
            if suffix == ".jsonl"
            else "csv"
        )

        output_dir = Path(target_config.get("output_dir", settings.output_dir))
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"{task_db_id}_desensitized{suffix}"

        total_rows = _count_rows(path, file_type)
        rows = _iter_rows(path, file_type)
        try:
            header = next(rows)
        except StopIteration:
            _update_task_record(
                task_db_id,
                status=TaskStatus.COMPLETED,
                progress=1.0,
                message="completed (no rows)",
                input_rows=0,
                output_rows=0,
                masked_fields=0,
                completed_at=datetime.utcnow(),
            )
            redis_client.publish(
                f"task_progress:{self.request.id}",
                json.dumps({"current": 0, "total": 0, "message": "completed"}),
            )
            return {"current": 0, "total": 0, "message": "completed"}

        header_cells = ["" if cell is None else str(cell) for cell in header]

        # Build maskers: prefer explicit rules from task config, fall back to auto-detect
        rules_list = rules_config.get("rules", [])
        if rules_list:
            from masking.engine import MaskingEngine

            engine = MaskingEngine.from_rules_config(rules_list)
            # Create a mapping from column index to masking rule
            maskers = []
            for cell in header_cells:
                rule = engine.get_rule(cell)
                maskers.append(rule)
            use_engine = True
        else:
            # Fall back to existing keyword-based auto-detection
            maskers = [_select_masker(cell) for cell in header_cells]
            use_engine = False

        masked_field_count = sum(1 for m in maskers if m is not None)
        data_total = max(total_rows - 1, 0)

        if file_type == "csv":
            with output_path.open("w", newline="") as handle:
                writer = csv.writer(handle)
                writer.writerow(header_cells)
                for index, row in enumerate(rows, start=1):
                    if use_engine:
                        # Build a dict row, apply engine, extract values
                        row_dict = {}
                        for i, cell in enumerate(row):
                            col = (
                                header_cells[i] if i < len(header_cells) else f"col_{i}"
                            )
                            row_dict[col] = "" if cell is None else str(cell)
                        masked_dict = engine.apply_to_row(row_dict)
                        masked_row = [masked_dict.get(col, "") for col in header_cells]
                    else:
                        masked_row = []
                        for i, cell in enumerate(row):
                            masker = maskers[i] if i < len(maskers) else None
                            masked_row.append(_apply_mask(cell, masker))
                    writer.writerow(masked_row)

                    if data_total and (index == data_total or index % 10 == 0):
                        progress = index / data_total if data_total else 1.0
                        message = f"Desensitizing row {index}/{data_total}"
                        self.update_state(
                            state="PROGRESS",
                            meta={
                                "current": index,
                                "total": data_total,
                                "message": message,
                            },
                        )
                        redis_client.publish(
                            f"task_progress:{self.request.id}",
                            json.dumps(
                                {
                                    "current": index,
                                    "total": data_total,
                                    "message": message,
                                }
                            ),
                        )
                        _update_task_record(
                            task_db_id,
                            progress=progress,
                            message=message,
                        )
        else:
            workbook = Workbook(write_only=True)
            sheet = workbook.create_sheet()
            sheet.append(header_cells)
            for index, row in enumerate(rows, start=1):
                if use_engine:
                    row_dict = {}
                    for i, cell in enumerate(row):
                        col = header_cells[i] if i < len(header_cells) else f"col_{i}"
                        row_dict[col] = "" if cell is None else str(cell)
                    masked_dict = engine.apply_to_row(row_dict)
                    masked_row = [masked_dict.get(col, "") for col in header_cells]
                else:
                    masked_row = []
                    for i, cell in enumerate(row):
                        masker = maskers[i] if i < len(maskers) else None
                        masked_row.append(_apply_mask(cell, masker))
                sheet.append(masked_row)

                if data_total and (index == data_total or index % 10 == 0):
                    progress = index / data_total if data_total else 1.0
                    message = f"Desensitizing row {index}/{data_total}"
                    self.update_state(
                        state="PROGRESS",
                        meta={
                            "current": index,
                            "total": data_total,
                            "message": message,
                        },
                    )
                    redis_client.publish(
                        f"task_progress:{self.request.id}",
                        json.dumps(
                            {"current": index, "total": data_total, "message": message}
                        ),
                    )
                    _update_task_record(
                        task_db_id,
                        progress=progress,
                        message=message,
                    )
            workbook.save(output_path)
            workbook.close()

        # Mark completed
        _update_task_record(
            task_db_id,
            status=TaskStatus.COMPLETED,
            progress=1.0,
            message="completed",
            input_rows=data_total,
            output_rows=data_total,
            masked_fields=masked_field_count,
            completed_at=datetime.utcnow(),
        )

        redis_client.publish(
            f"task_progress:{self.request.id}",
            json.dumps(
                {"current": data_total, "total": data_total, "message": "completed"}
            ),
        )

        return {
            "current": data_total,
            "total": data_total,
            "message": "completed",
            "output_file": output_path.name,
            "input_rows": data_total,
            "output_rows": data_total,
            "masked_fields": masked_field_count,
        }

    except Exception as exc:
        _update_task_record(
            task_db_id,
            status=TaskStatus.FAILED,
            error_detail=str(exc),
        )
        redis_client.publish(
            f"task_progress:{self.request.id}",
            json.dumps({"current": 0, "total": 0, "message": f"failed: {exc}"}),
        )
        raise
